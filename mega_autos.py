import os
import openpyxl
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import tkinter as tk

# Variables globales
lista_vehiculos = []

# Función para limpiar los campos del formulario
def Eliminar_Campos():
    entry_codigo.delete(0, END)
    entry_marca.delete(0, END)
    entry_modelo.delete(0, END)
    entry_precio.delete(0, END)
    entry_kilometraje.delete(0, END)

# Función para guardar vehículos en el archivo Excel
def guardar_vehiculos():
    codigo = entry_codigo.get()
    marca = entry_marca.get()
    modelo = entry_modelo.get()
    precio = entry_precio.get()
    kilometraje = entry_kilometraje.get()

    vehiculo = [codigo, marca, modelo, precio, kilometraje]
    lista_vehiculos.append(vehiculo)

    libro = openpyxl.load_workbook('vehiculos.xlsx')
    hoja = libro.active

    for vehiculo in lista_vehiculos:
        hoja.append(vehiculo)

    libro.save('vehiculos.xlsx')
    Eliminar_Campos()
    messagebox.showinfo('Agregar vehículo', 'Vehículo agregado con éxito')

# Función para eliminar vehículos del archivo Excel
def eliminar_vehiculos():
    codigo = entry_codigo.get()
    lista_vehiculos_eliminar = [vehiculo for vehiculo in lista_vehiculos if vehiculo[0] == codigo]

    if len(lista_vehiculos_eliminar) > 0:
        libro = openpyxl.load_workbook('vehiculos.xlsx')
        hoja = libro.active

        for row in hoja.iter_rows():
            for cell in row:
                if cell.value == codigo:
                    hoja.delete_rows(cell.row, 1)
                    break

        libro.save('vehiculos.xlsx')
        Eliminar_Campos()
        messagebox.showinfo('Eliminar vehículo', 'Vehículo eliminado con éxito')
    else:
        messagebox.showinfo('Eliminar vehículo', 'No se encontró el vehículo con el código ingresado')

# Función para listar vehículos del archivo Excel
def listar_vehiculos():
    codigo = entry_codigo.get()
    lista_vehiculos_buscar = [vehiculo for vehiculo in lista_vehiculos if vehiculo[0] == codigo]

    if len(lista_vehiculos_buscar) > 0:
        for vehiculo in lista_vehiculos_buscar:
            messagebox.showinfo('Listar vehículos', 'Código: {}, Marca: {}, Modelo: {}, Precio: {}, Kilometraje: {}'.format(vehiculo[0], vehiculo[1], vehiculo[2], vehiculo[3], vehiculo[4]))
    else:
        messagebox.showinfo('Listar vehículos', 'No se encontró el vehículo con el código ingresado')

# Crear interfaz de usuario con TKinter
ventana = tk.Tk()
ventana.title("Gestión de vehículos")

frame_botones = tk.Frame(ventana)
frame_botones.pack()

button_agregar = Button(frame_botones, text='Agregar', command=guardar_vehiculos)
button_agregar.pack(side=LEFT)

button_eliminar = Button(frame_botones, text='Eliminar', command=eliminar_vehiculos)
button_eliminar.pack(side=LEFT)

button_listar = Button(frame_botones, text='Listar', command=listar_vehiculos)
button_listar.pack(side=LEFT)

frame_ingreso = tk.Frame(ventana)
frame_ingreso.pack()

label_codigo = tk.Label(frame_ingreso, text="Código:")
label_codigo.pack(side=tk.LEFT  )
entry_codigo = tk.Entry(frame_ingreso)
entry_codigo.pack(side=tk.LEFT)

label_marca = tk.Label(frame_ingreso, text="Marca:")
label_marca.pack(side=tk.LEFT)
entry_marca = tk.Entry(frame_ingreso)
entry_marca.pack(side=tk.LEFT)

label_modelo = tk.Label(frame_ingreso, text="Modelo:")
label_modelo.pack(side=tk.LEFT)
entry_modelo = tk.Entry(frame_ingreso)
entry_modelo.pack(side=tk.LEFT)

label_precio = tk.Label(frame_ingreso, text="Precio:")
label_precio.pack(side=tk.LEFT)
entry_precio = tk.Entry(frame_ingreso)
entry_precio.pack(side=tk.LEFT)

label_kilometraje = tk.Label(frame_ingreso, text="Kilometraje:")
label_kilometraje.pack(side=tk.LEFT)
entry_kilometraje = tk.Entry(frame_ingreso)
entry_kilometraje.pack(side=tk.LEFT)

ventana.mainloop()